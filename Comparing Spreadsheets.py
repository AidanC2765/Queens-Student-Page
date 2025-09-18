import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from tkinter import Tk, filedialog

def select_files():
    root = Tk()
    root.withdraw()
    root.call('wm', 'attributes', '.', '-topmost', True)

    file_paths = filedialog.askopenfilenames(
        title="Select Two Excel Files",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    return list(file_paths)

def compare_model_numbers_only(file1, file2, output_path, key_column="Model_Number"):
    # Load both files
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Check if the key column exists
    if key_column not in df1.columns or key_column not in df2.columns:
        print(f"'{key_column}' column not found in both files.")
        return

    # Get sets of model numbers
    models1 = set(df1[key_column].dropna())
    models2 = set(df2[key_column].dropna())

    only_in_file1 = models1 - models2
    only_in_file2 = models2 - models1

    # Prepare Excel output
    wb = Workbook()
    ws = wb.active
    ws.title = "Model_Number Differences"

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True)

    ws.append(["Model_Number", "Status"])

    # Add rows only in file 1
    for model in sorted(only_in_file1):
        ws.append([model, "Only in File 1"])
        ws.cell(row=ws.max_row, column=1).fill = yellow_fill
        ws.cell(row=ws.max_row, column=1).alignment = wrap_alignment
        ws.cell(row=ws.max_row, column=2).alignment = wrap_alignment

    # Add rows only in file 2
    for model in sorted(only_in_file2):
        ws.append([model, "Only in File 2"])
        ws.cell(row=ws.max_row, column=1).fill = yellow_fill
        ws.cell(row=ws.max_row, column=1).alignment = wrap_alignment
        ws.cell(row=ws.max_row, column=2).alignment = wrap_alignment

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get column name
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save result
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Done! Differences saved in:\n{output_path}")

# --- Main Execution ---
if __name__ == "__main__":
    files = select_files()
    if len(files) != 2:
        print("⚠️ Please select exactly two Excel files.")
    else:
        output_file_path = r"Enter your file path location here"
        compare_model_numbers_only(files[0], files[1], output_file_path, key_column="Model_Number")

